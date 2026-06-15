
"""
App Streamlit - Comparador de Buques ISAOSA

Cómo ejecutar:
    pip install streamlit pandas openpyxl rapidfuzz
    streamlit run app_streamlit_isaosa.py

Flujo:
1. Carga ISAOSA, fuente azul y/o fuente naranja.
2. Lee, normaliza y filtra fuentes externas desde la fecha actual en adelante.
3. Compara externos contra ISAOSA por buque/viaje.
4. Separa resultados en:
   - Buques nuevos
   - Cambios detectados
   - Advertencias / Sin fecha
   - Resumen
5. Permite marcar decisiones:
   - Buques nuevos: AGREGAR / IGNORAR / PENDIENTE
   - Cambios: ACEPTAR / IGNORAR / PENDIENTE
6. Genera:
   - Reporte Excel
   - ISAOSA actualizado descargable, conservando el archivo original como base
"""

from __future__ import annotations

import io
import re
import unicodedata
from copy import copy
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process


# ============================================================
# CONFIGURACIÓN GENERAL
# ============================================================

APP_TITLE = "Comparador de Buques ISAOSA"
DEFAULT_ISAOSA_SHEET = "LINE-UP"
DEFAULT_BLUE_SHEET = "Hoja1"
DEFAULT_ORANGE_SHEETS = "AUTO"

SOURCE_ISAOSA = "ISAOSA"
SOURCE_BLUE = "AZUL"
SOURCE_ORANGE = "NARANJA"
SOURCE_EXTERNAL = "EXTERNO"

BLOCK_DATE_GAP_DAYS = 40
MAX_MATCH_DATE_DAYS = 60

ORANGE_ACTIVE_SHEETS = [
    "MANZANILLO",
    "TOPOLOBAMPO",
    "GUAYMAS",
    "VERACRUZ",
    "TAMPICO",
    "COATZACOALCOS",
    "ALTAMIRA",
    "PROGRESO",
    "TUXPAN",
]

ORANGE_EXCLUDE_SHEET_KEYWORDS = [
    "HISTORICO",
    "RESUMEN",
    "TOTAL",
    "DINAMICA",
]

DISPLAY_COLUMNS = [
    "vessel",
    "arrival_date",
    "tonnage",
    "cargo",
    "discharge_port",
    "importer",
    "supplier",
    "loading_port",
    "country_origin",
    "terminal",
    "pier",
    "etb",
    "etd",
    "status",
    "source",
    "source_file",
    "source_sheet",
    "source_row",
]

WORK_COLUMNS = [
    "vessel_clean",
    "cargo_clean",
    "discharge_port_clean",
    "importer_clean",
    "supplier_clean",
    "loading_port_clean",
    "country_origin_clean",
    "arrival_date_clean",
    "tonnage_clean",
]

FIELD_ALIASES = {
    "vessel": ["BARCO", "VESSEL", "VESSEL NAME", "MV", "M/V"],
    "arrival_date": ["ARRIVAL DATE", "ETA", "FECHA", "FECHA LLEGADA"],
    "tonnage": ["TONNAGE", "MT X VESSEL", "MTS", "MT", "TONELAJE"],
    "cargo": ["CARGO", "PRODUCT", "PRODUCTO", "CARGA"],
    "discharge_port": ["DISCH.PORT", "DISCH PORT", "DISCH. PORT", "DISPORT", "DISCHARGE PORT", "PUERTO", "PUERTO DESCARGA"],
    "importer": ["IMPORTER", "RECEIVER", "CLIENTE", "IMPORTADOR"],
    "supplier": ["TRADER/SUPPLIER", "TRADER/ SUPPLIER", "TRADER / SUPPLIER", "SHIPPER", "SUPPLIER", "PROVEEDOR"],
    "loading_port": ["ORIGEN", "LOADING PORT", "LOADPOART", "LOADPORT", "PUERTO ORIGEN"],
    "country_origin": ["COUNTRY OF ORIGIN", "ORIGIN COUNTRY", "PAIS ORIGEN", "PAÍS ORIGEN"],
    "terminal": ["TERMINAL"],
    "pier": ["PIER", "MUELLE"],
    "etb": ["ETB"],
    "etd": ["ETD"],
    "status": ["STATUS", "ESTATUS"],
}

UPDATE_FIELDS = [
    "arrival_date",
    "tonnage",
    "cargo",
    "discharge_port",
    "importer",
    "supplier",
    "loading_port",
    "country_origin",
    "terminal",
    "pier",
    "etb",
    "etd",
    "status",
]

NEW_VESSEL_FIELDS = [
    "vessel",
    "arrival_date",
    "tonnage",
    "cargo",
    "discharge_port",
    "importer",
    "supplier",
    "loading_port",
    "country_origin",
    "terminal",
    "pier",
    "etb",
    "etd",
    "status",
]

UNKNOWN_VALUES = {
    "",
    "TBC",
    "T B C",
    "T.B.C",
    "TBD",
    "T B D",
    "POR CONFIRMAR",
    "PENDIENTE",
    "SIN DEFINIR",
    "SIN DATO",
    "N/A",
    "NA",
    "NONE",
    "NULL",
}

INVALID_VESSEL_NAMES = {
    "A VER",
    "PRUEBA",
    "TEST",
    "SIN BUQUE",
    "SIN NOMBRE",
    "N A",
    "NA",
    "TBC",
    "TBD",
    "PENDIENTE",
    "POR CONFIRMAR",
}

PRODUCT_EQUIVALENCES = {
    "SAM GRAN": "SAM GRANULAR",
    "SAM GRANULAR": "SAM GRANULAR",
    "SAM STD": "SAM STD",
    "KCL GRANULAR": "KCL",
    "GMOP": "KCL",
    "MOP": "KCL",
    "UREA GRANULAR": "UREA GRANULAR",
    "GRANULAR 43 0 0": "UREA GRANULAR",
    "UREA 43 0 0 2S": "UREA AZUFRE",
    "UREA 43 0 0 2 5S": "UREA AZUFRE",
    "NS 43 0 0 2S": "UREA AZUFRE",
    "NS 43 0 0 2 5S": "UREA AZUFRE",
    "NK 20 0 4 19S": "NK 20-0-4+19S",
    "NK 20 0 4 19 S": "NK 20-0-4+19S",
}

PORT_EQUIVALENCES = {
    "MNZ": "MANZANILLO",
    "MANZ": "MANZANILLO",
    "MANZANILLO COLIMA": "MANZANILLO",
    "TOPO": "TOPOLOBAMPO",
    "TOPOLOB": "TOPOLOBAMPO",
    "VER": "VERACRUZ",
}

IMPORTER_EQUIVALENCES = {
    "GOMEZ": "FERT GOMEZ",
    "FERT GOMEZ": "FERT GOMEZ",
    "SOL NUTRIENTES": "SOL NUTRIENTES",
    "FERCO": "FERCO ATLANTICA",
    "FERCO ATLANTICA": "FERCO ATLANTICA",
    "FERT NUTRIENTES": "FERT NUTRIENTES",
}

SUPPLIER_EQUIVALENCES = {
    "LDG": "LDG TRADING",
    "LDG TRADING": "LDG TRADING",
    "GF TRADING": "GF TRADING",
    "GF TRADING AMEROPA": "GF TRADING",
    "MOCROSOURCE": "MICROSOURCE",
    "MICROSOURCE": "MICROSOURCE",
}

ORIGIN_EQUIVALENCES = {
    "CHINA": "CHINA",
    "TIANJIN CHINA": "CHINA",
    "RIZHAO CHINA": "CHINA",
    "BAYUQUAN CHINA": "CHINA",
    "QINZHOU CHINA": "CHINA",
    "VENEZUELA": "VENEZUELA",
    "JOSE VENEZUELA": "VENEZUELA",
    "ST PETERSBURG RUSSIA": "RUSSIA",
    "SAINT PETERSBURG RUSSIA": "RUSSIA",
    "UST LUGA RUSSIA": "RUSSIA",
}

ALIAS_TO_FIELD = {}
for standard, aliases in FIELD_ALIASES.items():
    for alias in aliases:
        ALIAS_TO_FIELD[alias] = standard


@dataclass
class ReadResult:
    data: pd.DataFrame
    warnings: list[dict[str, Any]]
    meta: dict[str, Any]


# ============================================================
# NORMALIZACIÓN
# ============================================================

def strip_accents(value: Any) -> str:
    text = str(value)
    return "".join(
        char for char in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(char)
    )


def clean_header_name(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = strip_accents(value)
    text = text.replace("\n", " ").replace("\r", " ")
    text = text.replace('"', "").replace("'", "")
    text = re.sub(r"[\u200b\u200c\u200d\ufeff]", "", text)
    text = re.sub(r"\s+", " ", text).strip().upper()
    return text


def header_key(value: Any) -> str:
    text = clean_header_name(value)
    text = text.replace("&", " AND ")
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def basic_normalize(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = strip_accents(value).upper()
    text = re.sub(r"[\u200b\u200c\u200d\ufeff]", "", text)
    text = re.sub(r"^(M\s*/\s*V|M\.?\s*V\.?|MV)\s+", "", text.strip())
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def is_unknown(value: Any) -> bool:
    key = basic_normalize(value)
    return key in UNKNOWN_VALUES


def normalize_text(value: Any) -> str:
    key = basic_normalize(value)
    return "" if key in UNKNOWN_VALUES else key


def normalize_date(value: Any) -> pd.Timestamp | pd.NaT:
    if is_unknown(value):
        return pd.NaT
    if isinstance(value, pd.Timestamp):
        return value.normalize()
    if isinstance(value, datetime):
        return pd.Timestamp(value).normalize()
    if isinstance(value, date):
        return pd.Timestamp(value).normalize()

    text = str(value).strip()
    if not text:
        return pd.NaT

    # Excel puede traer serial numérico
    if re.fullmatch(r"\d+(\.\d+)?", text):
        try:
            number = float(text)
            if 20000 <= number <= 80000:
                parsed = pd.to_datetime(number, unit="D", origin="1899-12-30", errors="coerce")
                if not pd.isna(parsed):
                    return pd.Timestamp(parsed).normalize()
        except Exception:
            pass

    if re.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$", text):
        parsed = pd.to_datetime(text, errors="coerce", dayfirst=False)
        if not pd.isna(parsed):
            return pd.Timestamp(parsed).normalize()

    parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        parsed = pd.to_datetime(text, errors="coerce", dayfirst=False)
    if pd.isna(parsed):
        return pd.NaT
    return pd.Timestamp(parsed).normalize()


def normalize_number(value: Any) -> float | None:
    if is_unknown(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = strip_accents(value).upper()
    text = text.replace(",", "").replace("$", "")
    text = re.sub(r"\b(MT|MTS|TON|TONS|TM)\b", "", text)
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group())
    except ValueError:
        return None


def apply_equivalence(value: Any, mapping: dict[str, str]) -> str:
    key = basic_normalize(value)
    if not key or key in UNKNOWN_VALUES:
        return ""
    return mapping.get(key, key)


def normalize_product(value: Any) -> str:
    key = basic_normalize(value)
    if not key or key in UNKNOWN_VALUES:
        return ""
    key = re.sub(r"(\d+)\s*[- ]\s*(\d+)\s*[- ]\s*(\d+)", r"\1 \2 \3", key)
    key = re.sub(r"\+", " ", key)
    key = re.sub(r"\s+", " ", key).strip()
    return PRODUCT_EQUIVALENCES.get(key, key)


def normalize_port(value: Any) -> str:
    return apply_equivalence(value, PORT_EQUIVALENCES)


def normalize_importer(value: Any) -> str:
    return apply_equivalence(value, IMPORTER_EQUIVALENCES)


def normalize_supplier(value: Any) -> str:
    return apply_equivalence(value, SUPPLIER_EQUIVALENCES)


def normalize_origin(value: Any) -> str:
    return apply_equivalence(value, ORIGIN_EQUIVALENCES)


def safe_display(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, pd.Timestamp):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value)


def format_date_for_excel(value: Any) -> date | Any:
    parsed = normalize_date(value)
    if not pd.isna(parsed):
        return parsed.date()
    return value


def is_invalid_vessel_name(value: Any) -> bool:
    clean = normalize_text(value)
    if not clean:
        return True
    if clean in INVALID_VESSEL_NAMES:
        return True
    if len(clean) < 3:
        return True
    if re.fullmatch(r"\d+", clean):
        return True
    if not re.search(r"[A-Z]", clean):
        return True
    alpha_count = len(re.findall(r"[A-Z]", clean))
    return alpha_count < 2


def warning_row(nivel: str, tipo: str, archivo: str = "", hoja: str = "", fila: Any = "", detalle: str = "") -> dict[str, Any]:
    return {
        "nivel": nivel,
        "tipo": tipo,
        "archivo": Path(str(archivo)).name if archivo else "",
        "hoja": hoja,
        "fila": fila,
        "detalle": detalle,
    }


# ============================================================
# LECTURA DE EXCEL
# ============================================================

def uploaded_name(uploaded_file) -> str:
    return uploaded_file.name if uploaded_file is not None else ""


def uploaded_bytes(uploaded_file) -> bytes:
    uploaded_file.seek(0)
    return uploaded_file.read()


def workbook_from_uploaded(uploaded_file, data_only: bool = False):
    return load_workbook(io.BytesIO(uploaded_bytes(uploaded_file)), data_only=data_only)


def sheet_names_from_uploaded(uploaded_file) -> list[str]:
    wb = workbook_from_uploaded(uploaded_file, data_only=True)
    return wb.sheetnames


def detect_header_row(uploaded_file, sheet_name: str, max_rows: int = 80) -> tuple[int, list[str], float]:
    data = uploaded_bytes(uploaded_file)
    preview = pd.read_excel(
        io.BytesIO(data),
        sheet_name=sheet_name,
        header=None,
        nrows=max_rows,
        dtype=object,
        engine="openpyxl",
    )
    alias_keys = {header_key(alias) for alias in ALIAS_TO_FIELD}
    best_row = None
    best_score = 0
    best_headers: list[str] = []

    for idx, row in preview.iterrows():
        values = [value for value in row.tolist() if clean_header_name(value)]
        keys = [header_key(value) for value in values]
        exact = len(set(keys) & alias_keys)
        fuzzy = 0
        for key in keys:
            if not key:
                continue
            match = process.extractOne(key, alias_keys, scorer=fuzz.ratio)
            if match and match[1] >= 92:
                fuzzy += 1
        density = min(len(values), 20) * 0.05
        score = exact * 2 + fuzzy + density
        if score > best_score:
            best_row = int(idx)
            best_score = score
            best_headers = [clean_header_name(value) for value in values]

    if best_row is None or best_score < 2:
        raise ValueError("No se pudo detectar una fila de encabezados confiable.")

    return best_row, best_headers, round(float(best_score), 2)


def map_columns(columns) -> tuple[list[str], list[str]]:
    used: dict[str, int] = {}
    mapped: list[str] = []
    unmapped: list[str] = []

    alias_keys = {header_key(alias): field for alias, field in ALIAS_TO_FIELD.items()}
    alias_lookup = list(alias_keys.keys())

    for col in columns:
        clean = clean_header_name(col)
        key = header_key(clean)
        standard = alias_keys.get(key)

        if standard is None and key:
            match = process.extractOne(key, alias_lookup, scorer=fuzz.ratio)
            if match and match[1] >= 94:
                standard = alias_keys[match[0]]

        if standard is None:
            name = clean if clean else "COLUMNA VACIA"
            unmapped.append(name)
            standard = f"extra_{len(unmapped)}_{header_key(name).lower().replace(' ', '_')}"

        if standard in used:
            used[standard] += 1
            standard = f"{standard}_{used[standard]}"
        else:
            used[standard] = 1

        mapped.append(standard)

    return mapped, unmapped


def resolve_sheet_name(uploaded_file, requested_sheet: str) -> tuple[str | None, dict[str, Any] | None]:
    names = sheet_names_from_uploaded(uploaded_file)
    if requested_sheet in names:
        return requested_sheet, None
    requested_key = header_key(requested_sheet)
    for sheet in names:
        if header_key(sheet) == requested_key:
            return sheet, warning_row(
                "ADVERTENCIA",
                "HOJA NORMALIZADA",
                uploaded_name(uploaded_file),
                requested_sheet,
                "",
                f"Se usó la hoja real '{sheet}'.",
            )
    return None, warning_row(
        "ERROR",
        "HOJA NO ENCONTRADA",
        uploaded_name(uploaded_file),
        requested_sheet,
        "",
        f"No existe la hoja '{requested_sheet}'. Hojas disponibles: {', '.join(names)}",
    )


def read_source_sheet(uploaded_file, sheet_name: str, source_name: str) -> ReadResult:
    warnings: list[dict[str, Any]] = []
    if uploaded_file is None:
        return ReadResult(empty_df(), warnings, {})

    real_sheet, sheet_warning = resolve_sheet_name(uploaded_file, sheet_name)
    if sheet_warning:
        warnings.append(sheet_warning)
    if not real_sheet:
        return ReadResult(empty_df(), warnings, {})

    try:
        header_row, headers_found, score = detect_header_row(uploaded_file, real_sheet)
        data = uploaded_bytes(uploaded_file)
        raw = pd.read_excel(
            io.BytesIO(data),
            sheet_name=real_sheet,
            header=header_row,
            dtype=object,
            engine="openpyxl",
        )
    except Exception as exc:
        warnings.append(warning_row(
            "ERROR",
            "ERROR DE LECTURA",
            uploaded_name(uploaded_file),
            real_sheet,
            "",
            str(exc),
        ))
        return ReadResult(empty_df(), warnings, {})

    original_columns = [clean_header_name(col) for col in raw.columns]
    mapped_columns, unmapped = map_columns(raw.columns)
    raw.columns = mapped_columns

    for col in DISPLAY_COLUMNS:
        if col not in raw.columns:
            raw[col] = None

    if unmapped:
        warnings.append(warning_row(
            "INFO",
            "COLUMNAS NO MAPEADAS",
            uploaded_name(uploaded_file),
            real_sheet,
            header_row + 1,
            ", ".join(unmapped[:30]),
        ))

    df = raw.copy()
    df["source"] = source_name
    df["source_file"] = uploaded_name(uploaded_file)
    df["source_sheet"] = real_sheet
    df["source_row"] = df.index + header_row + 2

    df["vessel_clean"] = df["vessel"].apply(normalize_text)
    df["cargo_clean"] = df["cargo"].apply(normalize_product)
    df["discharge_port_clean"] = df["discharge_port"].apply(normalize_port)
    df["importer_clean"] = df["importer"].apply(normalize_importer)
    df["supplier_clean"] = df["supplier"].apply(normalize_supplier)
    df["loading_port_clean"] = df["loading_port"].apply(normalize_origin)
    df["country_origin_clean"] = df["country_origin"].apply(normalize_origin)
    df["arrival_date_clean"] = df["arrival_date"].apply(normalize_date)
    df["tonnage_clean"] = df["tonnage"].apply(normalize_number)

    before = len(df)
    df = df[df["vessel_clean"] != ""].copy()
    removed_blank = before - len(df)

    invalid_mask = df["vessel"].apply(is_invalid_vessel_name)
    removed_invalid = int(invalid_mask.sum())
    df = df[~invalid_mask].copy()

    if removed_blank:
        warnings.append(warning_row(
            "INFO",
            "FILAS OMITIDAS",
            uploaded_name(uploaded_file),
            real_sheet,
            "",
            f"Se omitieron {removed_blank} filas sin buque.",
        ))
    if removed_invalid:
        warnings.append(warning_row(
            "INFO",
            "FILAS OMITIDAS",
            uploaded_name(uploaded_file),
            real_sheet,
            "",
            f"Se omitieron {removed_invalid} filas con nombre de buque inválido.",
        ))

    meta = {
        "archivo": uploaded_name(uploaded_file),
        "hoja": real_sheet,
        "source": source_name,
        "encabezado_fila": header_row + 1,
        "puntaje_encabezado": score,
        "encabezados_reales": ", ".join(original_columns),
        "encabezados_detectados": ", ".join(headers_found),
        "registros": len(df),
    }

    return ReadResult(df[DISPLAY_COLUMNS + WORK_COLUMNS].reset_index(drop=True), warnings, meta)


def empty_df() -> pd.DataFrame:
    return pd.DataFrame(columns=DISPLAY_COLUMNS + WORK_COLUMNS)


def resolve_orange_sheets(uploaded_file, sheets_value: str) -> tuple[list[str], list[str]]:
    names = sheet_names_from_uploaded(uploaded_file)
    requested = str(sheets_value or DEFAULT_ORANGE_SHEETS).strip()
    omitted: list[str] = []

    if requested.upper() == "AUTO":
        available_by_key = {header_key(sheet): sheet for sheet in names}
        active_keys = {header_key(sheet): sheet for sheet in ORANGE_ACTIVE_SHEETS}
        selected = []
        for key in active_keys:
            real_sheet = available_by_key.get(key)
            if real_sheet and not orange_sheet_is_excluded(real_sheet):
                selected.append(real_sheet)
        for sheet in names:
            if orange_sheet_is_excluded(sheet) or header_key(sheet) not in active_keys:
                omitted.append(sheet)
        return selected, omitted

    selected = [item.strip() for item in requested.split(",") if item.strip()]
    return selected, omitted


def orange_sheet_is_excluded(sheet_name: str) -> bool:
    key = header_key(sheet_name)
    return any(keyword in key for keyword in ORANGE_EXCLUDE_SHEET_KEYWORDS)


def load_orange(uploaded_file, sheets_value: str) -> ReadResult:
    warnings: list[dict[str, Any]] = []
    if uploaded_file is None:
        return ReadResult(empty_df(), warnings, {"registros": 0, "hojas_cargadas": []})

    try:
        sheets, omitted = resolve_orange_sheets(uploaded_file, sheets_value)
    except Exception as exc:
        warnings.append(warning_row("ERROR", "ARCHIVO NARANJA NO LEGIBLE", uploaded_name(uploaded_file), "", "", str(exc)))
        return ReadResult(empty_df(), warnings, {})

    frames = []
    loaded = []
    for sheet in sheets:
        result = read_source_sheet(uploaded_file, sheet, SOURCE_ORANGE)
        warnings.extend(result.warnings)
        if not result.data.empty:
            frames.append(result.data)
            loaded.append(result.meta.get("hoja", sheet))

    data = pd.concat(frames, ignore_index=True) if frames else empty_df()
    meta = {
        "archivo": uploaded_name(uploaded_file),
        "hojas_cargadas": loaded,
        "hojas_omitidas": omitted,
        "registros": len(data),
    }
    return ReadResult(data, warnings, meta)


# ============================================================
# FILTROS Y MATCHING
# ============================================================

def filter_external_current_onward(df: pd.DataFrame, source_name: str) -> tuple[pd.DataFrame, list[dict[str, Any]], dict[str, int]]:
    warnings: list[dict[str, Any]] = []
    if df is None or df.empty:
        return empty_df(), warnings, {
            "total_leido": 0,
            "excluidos_fecha_anterior": 0,
            "sin_fecha_valida": 0,
            "usados_para_comparar": 0,
        }

    today = pd.Timestamp.today().normalize()
    working = df.copy()
    invalid_mask = working["arrival_date_clean"].isna()
    past_mask = (~invalid_mask) & (working["arrival_date_clean"] < today)
    valid_mask = (~invalid_mask) & (working["arrival_date_clean"] >= today)

    for _, row in working[invalid_mask].iterrows():
        warnings.append(warning_row(
            "ADVERTENCIA",
            "SIN FECHA VÁLIDA",
            row.get("source_file", ""),
            row.get("source_sheet", ""),
            row.get("source_row", ""),
            f"Registro externo excluido: {safe_display(row.get('vessel', ''))}.",
        ))
    for _, row in working[past_mask].iterrows():
        warnings.append(warning_row(
            "INFO",
            "EXCLUIDO POR FECHA ANTERIOR A HOY",
            row.get("source_file", ""),
            row.get("source_sheet", ""),
            row.get("source_row", ""),
            f"{safe_display(row.get('vessel', ''))} | Fecha: {safe_display(row.get('arrival_date_clean', ''))}.",
        ))

    stats = {
        "total_leido": int(len(working)),
        "excluidos_fecha_anterior": int(past_mask.sum()),
        "sin_fecha_valida": int(invalid_mask.sum()),
        "usados_para_comparar": int(valid_mask.sum()),
    }
    warnings.append(warning_row(
        "INFO",
        "FILTRO DE FECHA",
        source_name,
        "",
        "",
        f"Leídos={stats['total_leido']} | excluidos anteriores={stats['excluidos_fecha_anterior']} | sin fecha={stats['sin_fecha_valida']} | usados={stats['usados_para_comparar']}",
    ))

    return working[valid_mask].copy().reset_index(drop=True), warnings, stats


def date_score(isaosa_date, ext_date) -> float:
    if pd.isna(isaosa_date) or pd.isna(ext_date):
        return 50
    days = abs((ext_date - isaosa_date).days)
    if days <= 3:
        return 100
    if days <= 10:
        return 90
    if days <= 30:
        return 75
    if days <= 60:
        return 45
    return 0


def tonnage_score(left, right) -> float:
    if left in (None, 0) or right in (None, 0):
        return 65
    try:
        pct = abs(float(right) - float(left)) / abs(float(left)) * 100
    except Exception:
        return 50
    if pct <= 3:
        return 100
    if pct <= 8:
        return 80
    if pct <= 20:
        return 55
    return 20


def text_similarity(left: str, right: str) -> float:
    if not left and not right:
        return 100
    if not left or not right:
        return 50
    return max(
        fuzz.token_sort_ratio(left, right),
        fuzz.token_set_ratio(left, right),
        fuzz.partial_ratio(left, right),
    )


def row_date(row) -> pd.Timestamp | pd.NaT:
    value = row.get("arrival_date_clean", pd.NaT)
    return value if not pd.isna(value) else pd.NaT


def block_date(block: dict[str, Any]):
    dates = [row_date(row) for row in block["rows"] if not pd.isna(row_date(row))]
    if not dates:
        return pd.NaT
    return min(dates)


def block_total_tonnage(block: dict[str, Any]) -> float | None:
    values = [row.get("tonnage_clean") for row in block["rows"] if row.get("tonnage_clean") is not None]
    return sum(values) if values else None


def block_values_text(block: dict[str, Any], field: str) -> str:
    values = []
    clean_field = f"{field}_clean"
    for row in block["rows"]:
        value = row.get(clean_field, "") if clean_field in row else normalize_text(row.get(field, ""))
        if value:
            values.append(value)
    return " ".join(sorted(set(values)))


def build_blocks(df: pd.DataFrame, source_label: str) -> list[dict[str, Any]]:
    blocks: list[dict[str, Any]] = []
    if df is None or df.empty:
        return blocks

    for vessel_clean, group in df[df["vessel_clean"] != ""].groupby("vessel_clean", sort=True):
        rows = [row for _, row in group.iterrows()]
        rows.sort(key=lambda row: (pd.isna(row_date(row)), row_date(row) if not pd.isna(row_date(row)) else pd.Timestamp.max))

        current = []
        last_date = pd.NaT
        for row in rows:
            current_date = row_date(row)
            should_split = False
            if current and not pd.isna(current_date) and not pd.isna(last_date):
                should_split = abs((current_date - last_date).days) > BLOCK_DATE_GAP_DAYS
            if should_split:
                blocks.append(make_block(current, source_label))
                current = []
            current.append(row)
            if not pd.isna(current_date):
                last_date = current_date

        if current:
            blocks.append(make_block(current, source_label))

    return blocks


def make_block(rows: list[pd.Series], source_label: str) -> dict[str, Any]:
    first = rows[0]
    return {
        "source": source_label,
        "vessel_clean": first.get("vessel_clean", ""),
        "vessel": safe_display(first.get("vessel", "")),
        "rows": rows,
    }


def classify_block_match(isaosa_block: dict[str, Any], external_block: dict[str, Any]) -> dict[str, Any]:
    vessel = fuzz.token_sort_ratio(external_block["vessel_clean"], isaosa_block["vessel_clean"])
    isa_date = block_date(isaosa_block)
    ext_date = block_date(external_block)

    if pd.isna(isa_date) or pd.isna(ext_date):
        days = None
        date_component = 55
    else:
        days = abs((ext_date - isa_date).days)
        if days <= 3:
            date_component = 100
        elif days <= 10:
            date_component = 92
        elif days <= 30:
            date_component = 78
        elif days <= 45:
            date_component = 62
        elif days <= 60:
            date_component = 35
        else:
            date_component = 0

    supplier = text_similarity(block_values_text(isaosa_block, "supplier"), block_values_text(external_block, "supplier"))
    port = text_similarity(block_values_text(isaosa_block, "discharge_port"), block_values_text(external_block, "discharge_port"))
    importer = text_similarity(block_values_text(isaosa_block, "importer"), block_values_text(external_block, "importer"))
    tonnage = tonnage_score(block_total_tonnage(isaosa_block), block_total_tonnage(external_block))

    total = vessel * 0.45 + date_component * 0.30 + supplier * 0.08 + port * 0.07 + tonnage * 0.05 + importer * 0.05

    reasons = [f"buque {round(vessel, 1)}%"]
    reasons.append("sin fecha suficiente" if days is None else f"fecha a {days} días")
    if supplier >= 80:
        reasons.append("proveedor parecido")
    if port >= 80:
        reasons.append("puerto parecido")
    if tonnage >= 80:
        reasons.append("tonelaje parecido")

    classification = "NO_MATCH"
    if vessel >= 92 and days is not None and days <= 10 and (supplier >= 70 or port >= 70 or tonnage >= 70):
        classification = "MATCH_SEGURO"
    elif vessel >= 90 and days is not None and days <= 30:
        classification = "MATCH_PROBABLE"
    elif vessel >= 92 and days is not None and days <= 45:
        classification = "REVISAR_MISMO_BUQUE"
    elif total >= 74 and vessel >= 88 and (days is None or days <= 60):
        classification = "REVISAR_MISMO_BUQUE"

    return {
        "classification": classification,
        "score": round(total, 2),
        "vessel": round(vessel, 2),
        "days": days,
        "supplier": round(supplier, 2),
        "port": round(port, 2),
        "importer": round(importer, 2),
        "tonnage": round(tonnage, 2),
        "reasons": reasons,
    }


def find_matching_block(isaosa_blocks: list[dict[str, Any]], external_block: dict[str, Any]) -> tuple[dict[str, Any] | None, dict[str, Any] | None]:
    candidates = []
    for block in isaosa_blocks:
        score = classify_block_match(block, external_block)
        if score["classification"] != "NO_MATCH":
            candidates.append((score["score"], block, score))

    candidates.sort(key=lambda item: item[0], reverse=True)
    if not candidates:
        return None, None
    return candidates[0][1], candidates[0][2]


def best_line_match(isaosa_rows: list[pd.Series], ext_row: pd.Series) -> tuple[pd.Series | None, float]:
    best = None
    best_score = -1.0
    for row in isaosa_rows:
        vessel = fuzz.token_sort_ratio(ext_row.get("vessel_clean", ""), row.get("vessel_clean", ""))
        d_score = date_score(row.get("arrival_date_clean", pd.NaT), ext_row.get("arrival_date_clean", pd.NaT))
        t_score = tonnage_score(row.get("tonnage_clean"), ext_row.get("tonnage_clean"))
        cargo = text_similarity(row.get("cargo_clean", ""), ext_row.get("cargo_clean", ""))
        port = text_similarity(row.get("discharge_port_clean", ""), ext_row.get("discharge_port_clean", ""))
        importer = text_similarity(row.get("importer_clean", ""), ext_row.get("importer_clean", ""))
        total = vessel * 0.35 + d_score * 0.25 + t_score * 0.15 + cargo * 0.10 + port * 0.10 + importer * 0.05
        if total > best_score:
            best_score = total
            best = row
    return best, round(best_score, 2)


def compare_dates(left, right):
    if pd.isna(right):
        return None
    if pd.isna(left):
        return "externo_tiene_dato"
    diff = int((right - left).days)
    return diff if diff != 0 else None


def compare_tonnage(left, right):
    if right is None:
        return None
    if left is None:
        return "externo_tiene_dato"
    if left == 0:
        return None
    pct = abs(right - left) / abs(left) * 100
    return round(pct, 2) if pct > 0.5 else None


def make_change_row(isa_row, ext_row, field, tipo, old, new, risk, action, reason, confidence, criterion) -> dict[str, Any]:
    return {
        "grupo_buque": normalize_text(ext_row.get("vessel", "")) or normalize_text(isa_row.get("vessel", "")),
        "buque_isaosa": safe_display(isa_row.get("vessel", "")),
        "buque_externo": safe_display(ext_row.get("vessel", "")),
        "campo": field,
        "valor_isaosa": safe_display(old),
        "valor_externo": safe_display(new),
        "tipo_cambio": tipo,
        "nivel_riesgo": risk,
        "confianza": confidence,
        "accion_sugerida": action,
        "motivo": reason,
        "criterio_match": criterion,
        "isaosa_source_sheet": safe_display(isa_row.get("source_sheet", "")),
        "isaosa_source_row": safe_display(isa_row.get("source_row", "")),
        "external_source_file": safe_display(ext_row.get("source_file", "")),
        "external_source_sheet": safe_display(ext_row.get("source_sheet", "")),
        "external_source_row": safe_display(ext_row.get("source_row", "")),
        "decision_usuario": "PENDIENTE",
        "row_id": "",
    }


def detect_line_changes(isa_row, ext_row, confidence: float, criterion: str) -> list[dict[str, Any]]:
    changes = []

    date_diff = compare_dates(isa_row.get("arrival_date_clean", pd.NaT), ext_row.get("arrival_date_clean", pd.NaT))
    if date_diff is not None:
        risk = "BAJO" if date_diff == "externo_tiene_dato" or abs(date_diff) <= 10 else "MEDIO"
        changes.append(make_change_row(
            isa_row,
            ext_row,
            "arrival_date",
            "CAMBIO FECHA",
            isa_row.get("arrival_date", ""),
            ext_row.get("arrival_date", ""),
            risk,
            "Actualizar fecha",
            "Diferencia de fecha" if date_diff != "externo_tiene_dato" else "ISAOSA no tiene fecha y externo sí",
            confidence,
            criterion,
        ))

    ton_diff = compare_tonnage(isa_row.get("tonnage_clean"), ext_row.get("tonnage_clean"))
    if ton_diff is not None and (ton_diff == "externo_tiene_dato" or ton_diff > 3):
        changes.append(make_change_row(
            isa_row,
            ext_row,
            "tonnage",
            "CAMBIO TONELAJE",
            isa_row.get("tonnage", ""),
            ext_row.get("tonnage", ""),
            "ALTO",
            "Revisar tonelaje",
            f"Diferencia aproximada {ton_diff}%" if ton_diff != "externo_tiene_dato" else "ISAOSA no tiene tonelaje y externo sí",
            confidence,
            criterion,
        ))

    text_fields = [
        ("cargo", "cargo_clean", "CAMBIO CARGA", "ALTO", "Revisar carga"),
        ("discharge_port", "discharge_port_clean", "CAMBIO PUERTO", "ALTO", "Revisar puerto"),
        ("importer", "importer_clean", "CAMBIO IMPORTADOR", "ALTO", "Revisar importador"),
        ("supplier", "supplier_clean", "CAMBIO PROVEEDOR", "MEDIO", "Validar proveedor"),
        ("loading_port", "loading_port_clean", "CAMBIO ORIGEN", "MEDIO", "Validar origen"),
        ("country_origin", "country_origin_clean", "CAMBIO PAÍS ORIGEN", "MEDIO", "Validar país origen"),
    ]

    for original, clean, tipo, risk, action in text_fields:
        ext_clean = ext_row.get(clean, "")
        isa_clean = isa_row.get(clean, "")
        if not ext_clean:
            continue
        if not isa_clean:
            changes.append(make_change_row(
                isa_row,
                ext_row,
                original,
                tipo,
                isa_row.get(original, ""),
                ext_row.get(original, ""),
                "MEDIO",
                "Completar dato faltante",
                "ISAOSA no tiene dato y externo sí",
                confidence,
                criterion,
            ))
            continue
        similarity = text_similarity(isa_clean, ext_clean)
        threshold = 88 if original in {"cargo", "discharge_port", "importer"} else 82
        if similarity < threshold:
            changes.append(make_change_row(
                isa_row,
                ext_row,
                original,
                tipo,
                isa_row.get(original, ""),
                ext_row.get(original, ""),
                risk,
                action,
                "El valor normalizado no coincide",
                confidence,
                criterion,
            ))

    return changes


def build_new_rows_from_block(external_block: dict[str, Any], score: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    rows = []
    for ext_row in external_block["rows"]:
        buque = safe_display(ext_row.get("vessel", ""))
        row = {
            "grupo_buque": normalize_text(buque),
            "buque": buque,
            "fecha": safe_display(ext_row.get("arrival_date_clean", ext_row.get("arrival_date", ""))),
            "tonelaje": safe_display(ext_row.get("tonnage", "")),
            "carga": safe_display(ext_row.get("cargo", "")),
            "puerto_descarga": safe_display(ext_row.get("discharge_port", "")),
            "importador": safe_display(ext_row.get("importer", "")),
            "proveedor": safe_display(ext_row.get("supplier", "")),
            "origen": safe_display(ext_row.get("loading_port", "")) or safe_display(ext_row.get("country_origin", "")),
            "fuente_externa": safe_display(ext_row.get("source", "")),
            "archivo_externo": safe_display(ext_row.get("source_file", "")),
            "hoja_externa": safe_display(ext_row.get("source_sheet", "")),
            "fila_externa": safe_display(ext_row.get("source_row", "")),
            "motivo": "No se encontró coincidencia confiable en ISAOSA",
            "decision_usuario": "PENDIENTE",
            "row_id": "",
        }

        for field in NEW_VESSEL_FIELDS:
            row[field] = ext_row.get(field, "")

        rows.append(row)
    return rows


def compare_all(isaosa_df: pd.DataFrame, external_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    new_rows: list[dict[str, Any]] = []
    change_rows: list[dict[str, Any]] = []

    isaosa_blocks = build_blocks(isaosa_df, SOURCE_ISAOSA)
    external_blocks = build_blocks(external_df, SOURCE_EXTERNAL)

    for ext_block in external_blocks:
        isa_block, score = find_matching_block(isaosa_blocks, ext_block)

        if isa_block is None or score is None or score["classification"] == "NO_MATCH":
            new_rows.extend(build_new_rows_from_block(ext_block, score))
            continue

        criterion = f"{score['classification']} | score={score['score']} | " + "; ".join(score.get("reasons", []))
        for ext_row in ext_block["rows"]:
            isa_row, line_score = best_line_match(isa_block["rows"], ext_row)
            if isa_row is None:
                new_rows.extend(build_new_rows_from_block({"rows": [ext_row]}, score))
                continue

            changes = detect_line_changes(isa_row, ext_row, line_score, criterion)
            change_rows.extend(changes)

    new_df = pd.DataFrame(new_rows)
    change_df = pd.DataFrame(change_rows)

    if not new_df.empty:
        new_df = new_df.sort_values(["grupo_buque", "fecha", "fila_externa"]).reset_index(drop=True)
        new_df["row_id"] = [f"N{i+1:05d}" for i in range(len(new_df))]

    if not change_df.empty:
        change_df = change_df.sort_values(["grupo_buque", "campo", "external_source_row"]).reset_index(drop=True)
        change_df["row_id"] = [f"C{i+1:05d}" for i in range(len(change_df))]

    return new_df, change_df


# ============================================================
# EXCEL DE SALIDA
# ============================================================

def df_to_excel_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, df in sheets.items():
            clean_name = name[:31]
            safe_df = df.copy() if df is not None else pd.DataFrame()
            for col in safe_df.columns:
                if pd.api.types.is_datetime64_any_dtype(safe_df[col]):
                    safe_df[col] = safe_df[col].dt.strftime("%d/%m/%Y")
            safe_df.to_excel(writer, index=False, sheet_name=clean_name)

        wb = writer.book
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.alignment = Alignment(horizontal="center")
            for column_cells in ws.columns:
                max_len = 10
                col_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    value = cell.value
                    if value is not None:
                        max_len = max(max_len, min(len(str(value)) + 2, 60))
                ws.column_dimensions[col_letter].width = max_len

    return output.getvalue()


def find_header_row_openpyxl(ws, max_rows: int = 80) -> int:
    alias_keys = {header_key(alias) for alias in ALIAS_TO_FIELD}
    best_row = 1
    best_score = 0

    for r in range(1, min(ws.max_row, max_rows) + 1):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        keys = [header_key(v) for v in values if header_key(v)]
        exact = len(set(keys) & alias_keys)
        fuzzy = 0
        for key in keys:
            match = process.extractOne(key, alias_keys, scorer=fuzz.ratio)
            if match and match[1] >= 92:
                fuzzy += 1
        score = exact * 2 + fuzzy
        if score > best_score:
            best_score = score
            best_row = r

    return best_row


def map_openpyxl_columns(ws, header_row: int) -> dict[str, int]:
    columns = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    mapped, _ = map_columns(columns)
    colmap: dict[str, int] = {}
    for idx, field in enumerate(mapped, start=1):
        if field not in colmap:
            colmap[field] = idx
    return colmap


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(source_row, col)
        target_cell = ws.cell(target_row, col)
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)


def write_value_to_cell(cell, field: str, value: Any) -> None:
    if field in {"arrival_date", "etb", "etd"}:
        parsed = normalize_date(value)
        if not pd.isna(parsed):
            cell.value = parsed.date()
            cell.number_format = "dd/mm/yyyy"
        else:
            cell.value = value
        return

    if field == "tonnage":
        number = normalize_number(value)
        cell.value = number if number is not None else value
        return

    cell.value = value


def create_updated_isaosa_bytes(
    isaosa_uploaded,
    isaosa_sheet: str,
    changes_df: pd.DataFrame,
    new_df: pd.DataFrame,
) -> bytes:
    original_bytes = uploaded_bytes(isaosa_uploaded)
    wb = load_workbook(io.BytesIO(original_bytes))
    if isaosa_sheet not in wb.sheetnames:
        # Intentar normalizar por nombre
        requested_key = header_key(isaosa_sheet)
        found = None
        for sheet in wb.sheetnames:
            if header_key(sheet) == requested_key:
                found = sheet
                break
        if not found:
            raise ValueError(f"No encontré la hoja ISAOSA '{isaosa_sheet}' en el archivo.")
        isaosa_sheet = found

    ws = wb[isaosa_sheet]
    header_row = find_header_row_openpyxl(ws)
    colmap = map_openpyxl_columns(ws, header_row)

    audit_rows: list[dict[str, Any]] = []

    # A) Aplicar cambios aceptados
    if changes_df is not None and not changes_df.empty:
        accepted = changes_df[changes_df["decision_usuario"].astype(str).str.upper().eq("ACEPTAR")]
        for _, row in accepted.iterrows():
            field = str(row.get("campo", "")).strip()
            if not field:
                continue
            target_col = colmap.get(field)
            target_row = row.get("isaosa_source_row", "")
            try:
                target_row = int(float(target_row))
            except Exception:
                audit_rows.append({
                    "accion": "ACTUALIZADO",
                    "buque": row.get("buque_isaosa", ""),
                    "campo": field,
                    "valor_anterior": row.get("valor_isaosa", ""),
                    "valor_nuevo": row.get("valor_externo", ""),
                    "fila_afectada": "",
                    "estado": "OMITIDO",
                    "detalle": "No se pudo identificar la fila ISAOSA.",
                })
                continue

            if not target_col:
                audit_rows.append({
                    "accion": "ACTUALIZADO",
                    "buque": row.get("buque_isaosa", ""),
                    "campo": field,
                    "valor_anterior": row.get("valor_isaosa", ""),
                    "valor_nuevo": row.get("valor_externo", ""),
                    "fila_afectada": target_row,
                    "estado": "OMITIDO",
                    "detalle": f"No existe columna mapeada para {field}.",
                })
                continue

            cell = ws.cell(target_row, target_col)
            previous = cell.value
            write_value_to_cell(cell, field, row.get("valor_externo", ""))
            audit_rows.append({
                "accion": "ACTUALIZADO",
                "buque": row.get("buque_isaosa", ""),
                "campo": field,
                "valor_anterior": previous,
                "valor_nuevo": row.get("valor_externo", ""),
                "fila_afectada": target_row,
                "estado": "APLICADO",
                "detalle": "Cambio aceptado aplicado.",
            })

    # B) Agregar buques nuevos
    if new_df is not None and not new_df.empty:
        to_add = new_df[new_df["decision_usuario"].astype(str).str.upper().eq("AGREGAR")]
        for _, row in to_add.iterrows():
            target_row = ws.max_row + 1
            copy_row_style(ws, max(header_row + 1, target_row - 1), target_row)

            for field in NEW_VESSEL_FIELDS:
                target_col = colmap.get(field)
                if not target_col:
                    audit_rows.append({
                        "accion": "AGREGADO",
                        "buque": row.get("buque", row.get("vessel", "")),
                        "campo": field,
                        "valor_anterior": "",
                        "valor_nuevo": row.get(field, ""),
                        "fila_afectada": target_row,
                        "estado": "OMITIDO",
                        "detalle": f"No existe columna mapeada para {field}.",
                    })
                    continue

                value = row.get(field, "")
                if field == "arrival_date" and (value is None or safe_display(value) == ""):
                    value = row.get("fecha", "")

                write_value_to_cell(ws.cell(target_row, target_col), field, value)

                audit_rows.append({
                    "accion": "AGREGADO",
                    "buque": row.get("buque", row.get("vessel", "")),
                    "campo": field,
                    "valor_anterior": "",
                    "valor_nuevo": safe_display(value),
                    "fila_afectada": target_row,
                    "estado": "APLICADO",
                    "detalle": "Buque nuevo agregado.",
                })

    # C) Auditoría
    audit_name = "Cambios aplicados"
    if audit_name in wb.sheetnames:
        del wb[audit_name]
    audit_ws = wb.create_sheet(audit_name)
    audit_columns = ["accion", "buque", "campo", "valor_anterior", "valor_nuevo", "fila_afectada", "estado", "detalle"]

    for c, col in enumerate(audit_columns, start=1):
        cell = audit_ws.cell(1, c, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for r, audit in enumerate(audit_rows, start=2):
        for c, col in enumerate(audit_columns, start=1):
            audit_ws.cell(r, c, audit.get(col, ""))

    audit_ws.freeze_panes = "A2"
    audit_ws.auto_filter.ref = audit_ws.dimensions
    for c in range(1, len(audit_columns) + 1):
        audit_ws.column_dimensions[get_column_letter(c)].width = 22

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ============================================================
# UI STREAMLIT
# ============================================================

def initialize_state():
    defaults = {
        "new_df": pd.DataFrame(),
        "changes_df": pd.DataFrame(),
        "warnings_df": pd.DataFrame(),
        "summary_df": pd.DataFrame(),
        "log_df": pd.DataFrame(),
        "report_bytes": None,
        "updated_bytes": None,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def add_log(logs: list[dict[str, Any]], etapa: str, detalle: str):
    logs.append({
        "momento": datetime.now().strftime("%H:%M:%S"),
        "etapa": etapa,
        "detalle": detalle,
    })


def render_grouped_editor(
    df: pd.DataFrame,
    key_prefix: str,
    decision_options: list[str],
    group_col: str = "grupo_buque",
) -> pd.DataFrame:
    if df is None or df.empty:
        st.info("No hay registros para mostrar.")
        return pd.DataFrame()

    edited_parts = []
    search = st.text_input("Buscar buque", key=f"{key_prefix}_search").strip().upper()

    working = df.copy()
    if search:
        mask = working[group_col].astype(str).str.upper().str.contains(search, na=False) | working.astype(str).apply(
            lambda row: row.str.upper().str.contains(search, na=False).any(), axis=1
        )
        working = working[mask].copy()

    groups = list(working.groupby(group_col, sort=True))
    st.caption(f"Mostrando {len(working)} registros en {len(groups)} grupo(s).")

    for group_name, group_df in groups:
        label = group_df.iloc[0].get("buque", "") or group_df.iloc[0].get("buque_externo", "") or group_name
        with st.expander(f"🚢 {label} · {len(group_df)} registro(s)", expanded=len(groups) <= 3):
            left, right = st.columns([1, 3])
            with left:
                bulk = st.selectbox(
                    "Marcar grupo como",
                    ["Sin cambio"] + decision_options,
                    key=f"{key_prefix}_bulk_{group_name}",
                )
                if bulk != "Sin cambio":
                    group_df = group_df.copy()
                    group_df["decision_usuario"] = bulk
            with right:
                edited = st.data_editor(
                    group_df,
                    key=f"{key_prefix}_editor_{group_name}",
                    hide_index=True,
                    use_container_width=True,
                    column_config={
                        "decision_usuario": st.column_config.SelectboxColumn(
                            "Decisión",
                            options=decision_options,
                            required=True,
                        )
                    },
                    disabled=[c for c in group_df.columns if c not in {"decision_usuario"}],
                )
                edited_parts.append(edited)

    if edited_parts:
        edited_all = pd.concat(edited_parts, ignore_index=True)
    else:
        edited_all = working.copy()

    # Mezclar editados filtrados con los que no se mostraron
    if search:
        shown_ids = set(edited_all.get("row_id", []))
        remaining = df[~df["row_id"].isin(shown_ids)].copy()
        final = pd.concat([remaining, edited_all], ignore_index=True)
        return final.sort_values([group_col, "row_id"]).reset_index(drop=True)

    return edited_all.sort_values([group_col, "row_id"]).reset_index(drop=True)


def build_summary(
    isaosa_result: ReadResult,
    blue_result: ReadResult,
    orange_result: ReadResult,
    blue_stats: dict[str, int],
    orange_stats: dict[str, int],
    new_df: pd.DataFrame,
    changes_df: pd.DataFrame,
    warnings_df: pd.DataFrame,
) -> pd.DataFrame:
    today = pd.Timestamp.today().normalize()
    rows = [
        {"seccion": "Fecha de corte", "detalle": f"Se comparan fuentes externas con fecha >= {today.strftime('%d/%m/%Y')}."},
        {"seccion": "ISAOSA", "detalle": f"Registros leídos: {len(isaosa_result.data)}."},
        {"seccion": "Fuente azul", "detalle": f"Leídos={blue_stats.get('total_leido', 0)} | usados={blue_stats.get('usados_para_comparar', 0)} | anteriores={blue_stats.get('excluidos_fecha_anterior', 0)} | sin fecha={blue_stats.get('sin_fecha_valida', 0)}."},
        {"seccion": "Fuente naranja", "detalle": f"Leídos={orange_stats.get('total_leido', 0)} | usados={orange_stats.get('usados_para_comparar', 0)} | anteriores={orange_stats.get('excluidos_fecha_anterior', 0)} | sin fecha={orange_stats.get('sin_fecha_valida', 0)}."},
        {"seccion": "Buques nuevos", "detalle": f"{len(new_df)} registro(s)."},
        {"seccion": "Cambios detectados", "detalle": f"{len(changes_df)} cambio(s)."},
        {"seccion": "Advertencias", "detalle": f"{len(warnings_df)} advertencia(s)."},
    ]

    if changes_df is not None and not changes_df.empty:
        for tipo, count in changes_df["tipo_cambio"].value_counts().items():
            rows.append({"seccion": "Cambios por tipo", "detalle": f"{tipo}: {count}"})

    return pd.DataFrame(rows)


def run_comparison(isaosa_file, blue_file, orange_file, isaosa_sheet: str, blue_sheet: str, orange_sheets: str):
    logs: list[dict[str, Any]] = []
    all_warnings: list[dict[str, Any]] = []

    add_log(logs, "Lectura", "Leyendo ISAOSA.")
    isaosa_result = read_source_sheet(isaosa_file, isaosa_sheet, SOURCE_ISAOSA)
    all_warnings.extend(isaosa_result.warnings)

    add_log(logs, "Lectura", "Leyendo fuente azul.")
    blue_result = read_source_sheet(blue_file, blue_sheet, SOURCE_BLUE) if blue_file is not None else ReadResult(empty_df(), [], {"registros": 0})
    all_warnings.extend(blue_result.warnings)

    add_log(logs, "Lectura", "Leyendo fuente naranja.")
    orange_result = load_orange(orange_file, orange_sheets) if orange_file is not None else ReadResult(empty_df(), [], {"registros": 0})
    all_warnings.extend(orange_result.warnings)

    add_log(logs, "Filtro", "Filtrando fuentes externas desde la fecha actual en adelante.")
    blue_filtered, blue_warnings, blue_stats = filter_external_current_onward(blue_result.data, "Fuente azul")
    orange_filtered, orange_warnings, orange_stats = filter_external_current_onward(orange_result.data, "Fuente naranja")
    all_warnings.extend(blue_warnings)
    all_warnings.extend(orange_warnings)

    external_df = pd.concat([blue_filtered, orange_filtered], ignore_index=True) if not blue_filtered.empty or not orange_filtered.empty else empty_df()

    add_log(logs, "Comparación", "Comparando fuentes externas contra ISAOSA.")
    new_df, changes_df = compare_all(isaosa_result.data, external_df)

    warnings_df = pd.DataFrame(all_warnings)
    summary_df = build_summary(isaosa_result, blue_result, orange_result, blue_stats, orange_stats, new_df, changes_df, warnings_df)
    log_df = pd.DataFrame(logs)

    st.session_state.new_df = new_df
    st.session_state.changes_df = changes_df
    st.session_state.warnings_df = warnings_df
    st.session_state.summary_df = summary_df
    st.session_state.log_df = log_df
    st.session_state.report_bytes = None
    st.session_state.updated_bytes = None


def main():
    st.set_page_config(page_title=APP_TITLE, page_icon="🚢", layout="wide")
    initialize_state()

    st.title("🚢 Comparador de Buques ISAOSA")
    st.caption("Versión Streamlit · resultados separados por pestañas · comparación desde la fecha actual en adelante")

    with st.sidebar:
        st.header("Archivos")
        isaosa_file = st.file_uploader("Base ISAOSA", type=["xlsx", "xlsm"], key="isaosa_file")
        blue_file = st.file_uploader("Fuente azul", type=["xlsx", "xlsm"], key="blue_file")
        orange_file = st.file_uploader("Fuente naranja", type=["xlsx", "xlsm"], key="orange_file")

        st.divider()
        st.header("Hojas")
        isaosa_sheet = st.text_input("Hoja ISAOSA", value=DEFAULT_ISAOSA_SHEET)
        blue_sheet = st.text_input("Hoja fuente azul", value=DEFAULT_BLUE_SHEET)
        orange_sheets = st.text_input("Hojas naranja", value=DEFAULT_ORANGE_SHEETS, help="Usa AUTO o separa nombres por coma.")

        st.divider()
        if st.button("🔎 Detectar cambios", type="primary", use_container_width=True):
            if isaosa_file is None:
                st.error("Carga primero la base ISAOSA.")
            elif blue_file is None and orange_file is None:
                st.error("Carga al menos una fuente externa: azul o naranja.")
            else:
                try:
                    with st.spinner("Procesando archivos..."):
                        run_comparison(isaosa_file, blue_file, orange_file, isaosa_sheet, blue_sheet, orange_sheets)
                    st.success("Comparación terminada.")
                except Exception as exc:
                    st.exception(exc)

    new_df = st.session_state.new_df
    changes_df = st.session_state.changes_df
    warnings_df = st.session_state.warnings_df
    summary_df = st.session_state.summary_df
    log_df = st.session_state.log_df

    # Métricas
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Buques nuevos", len(new_df) if new_df is not None else 0)
    m2.metric("Cambios detectados", len(changes_df) if changes_df is not None else 0)
    m3.metric("Advertencias", len(warnings_df) if warnings_df is not None else 0)
    m4.metric("Fecha de corte", pd.Timestamp.today().strftime("%d/%m/%Y"))

    tab_new, tab_changes, tab_warn, tab_summary, tab_exports = st.tabs([
        "🆕 Buques nuevos",
        "✏️ Cambios detectados",
        "⚠️ Advertencias / Sin fecha",
        "📊 Resumen / Log",
        "📦 Exportaciones",
    ])

    with tab_new:
        st.subheader("Buques nuevos")
        st.caption("Solo viajes externos que no fueron encontrados en ISAOSA. Marca AGREGAR para insertarlos en el ISAOSA actualizado.")
        edited_new = render_grouped_editor(new_df, "new", ["AGREGAR", "IGNORAR", "PENDIENTE"])
        if not edited_new.empty:
            st.session_state.new_df = edited_new

    with tab_changes:
        st.subheader("Cambios detectados")
        st.caption("Solo buques encontrados en ISAOSA que tienen diferencias reales. Marca ACEPTAR para aplicar el cambio.")
        edited_changes = render_grouped_editor(changes_df, "changes", ["ACEPTAR", "IGNORAR", "PENDIENTE"])
        if not edited_changes.empty:
            st.session_state.changes_df = edited_changes

    with tab_warn:
        st.subheader("Advertencias / Sin fecha")
        if warnings_df is None or warnings_df.empty:
            st.success("No hay advertencias.")
        else:
            st.dataframe(warnings_df, use_container_width=True, hide_index=True)

    with tab_summary:
        st.subheader("Resumen")
        if summary_df is None or summary_df.empty:
            st.info("Ejecuta la comparación para ver el resumen.")
        else:
            st.dataframe(summary_df, use_container_width=True, hide_index=True)

        st.subheader("Log")
        if log_df is None or log_df.empty:
            st.info("No hay log todavía.")
        else:
            st.dataframe(log_df, use_container_width=True, hide_index=True)

    with tab_exports:
        st.subheader("Exportaciones")

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("### Reporte Excel")
            if st.button("Generar reporte", use_container_width=True):
                sheets = {
                    "Buques nuevos": st.session_state.new_df,
                    "Cambios detectados": st.session_state.changes_df,
                    "Advertencias": st.session_state.warnings_df,
                    "Resumen": st.session_state.summary_df,
                    "Log": st.session_state.log_df,
                }
                st.session_state.report_bytes = df_to_excel_bytes(sheets)

            if st.session_state.report_bytes:
                st.download_button(
                    "Descargar reporte Excel",
                    data=st.session_state.report_bytes,
                    file_name=f"reporte_buques_isaosa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        with col2:
            st.markdown("### ISAOSA actualizado")
            st.caption("Genera un archivo nuevo descargable, tomando tu ISAOSA original como base. No modifica el archivo subido.")
            accepted_count = 0 if st.session_state.changes_df.empty else int(st.session_state.changes_df["decision_usuario"].astype(str).str.upper().eq("ACEPTAR").sum())
            add_count = 0 if st.session_state.new_df.empty else int(st.session_state.new_df["decision_usuario"].astype(str).str.upper().eq("AGREGAR").sum())
            st.info(f"Cambios a aplicar: {accepted_count} · Buques a agregar: {add_count}")

            if st.button("Generar ISAOSA actualizado", type="primary", use_container_width=True):
                if isaosa_file is None:
                    st.error("Carga la base ISAOSA.")
                elif accepted_count == 0 and add_count == 0:
                    st.warning("No hay cambios ACEPTAR ni buques AGREGAR.")
                else:
                    try:
                        st.session_state.updated_bytes = create_updated_isaosa_bytes(
                            isaosa_file,
                            isaosa_sheet,
                            st.session_state.changes_df,
                            st.session_state.new_df,
                        )
                        st.success("ISAOSA actualizado generado.")
                    except Exception as exc:
                        st.exception(exc)

            if st.session_state.updated_bytes:
                original_name = Path(uploaded_name(isaosa_file)).stem if isaosa_file else "ISAOSA"
                st.download_button(
                    "Descargar ISAOSA actualizado",
                    data=st.session_state.updated_bytes,
                    file_name=f"{original_name}_actualizado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

    st.divider()
    st.caption("Nota: el archivo ISAOSA actualizado se genera como copia modificada del original para conservar hojas, formatos y estructura general.")


if __name__ == "__main__":
    main()
